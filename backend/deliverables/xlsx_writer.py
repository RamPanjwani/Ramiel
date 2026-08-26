"""Excel Spreadsheet (.xlsx) Deliverable Generator for Ramiel.

Phase 5: Deliverable Generation.
Generates structured Excel workbooks (.xlsx) with computed tables, formulas,
and formatting from agent calculation results using openpyxl.
"""

from __future__ import annotations

import uuid
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


class XlsxWriter:
    """Generates formatted Microsoft Excel (.xlsx) workbooks."""

    def __init__(self, default_output_dir: str | Path = "data/uploads") -> None:
        self.output_dir = Path(default_output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        data: dict[str, Any],
        filename: str | None = None,
    ) -> str:
        """Generate an Excel workbook from structured tabular and calculation data.

        Args:
            data: Structured dictionary containing:
                - 'sheets': List of dicts with 'sheet_name' (str), 'headers' (list[str]), 'rows' (list[list[Any]])
                - 'title': Optional title (str)
            filename: Custom filename (defaults to generated UUID filename)

        Returns:
            The filesystem path to the generated .xlsx file.
        """
        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)  # type: ignore[arg-type]

        sheets_list = data.get("sheets", [])
        if not sheets_list:
            sheets_list = [
                {
                    "sheet_name": "Summary",
                    "headers": data.get("headers", ["Item", "Value"]),
                    "rows": data.get("rows", [["Status", "Completed"]]),
                }
            ]

        header_fill = PatternFill(
            start_color="161B1E", end_color="161B1E", fill_type="solid"
        )
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Arial", size=10)

        for sheet_info in sheets_list:
            sheet_name = sheet_info.get("sheet_name", "Sheet1")
            ws = wb.create_sheet(title=sheet_name)

            headers = sheet_info.get("headers", [])
            rows = sheet_info.get("rows", [])

            # Write header row
            if headers:
                ws.append(headers)
                for col_num in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font

            # Write data rows
            for row in rows:
                ws.append(row)

            # Apply styling & adjust column widths
            for row in ws.iter_rows(
                min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)
            ):
                for cell in row:
                    cell.font = data_font

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        out_filename = filename or f"workbook_{uuid.uuid4().hex[:6]}.xlsx"
        out_path = self.output_dir / out_filename
        wb.save(str(out_path))
        return str(out_path)
