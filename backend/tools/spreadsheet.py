"""Spreadsheet Operations Tool for Ramiel.

Phase 3: Tool Layer.
Provides Excel (.xlsx) spreadsheet inspection, reading, calculation formula extraction,
and writing using openpyxl and pandas.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


class SpreadsheetTool:
    """Tool for reading and modifying Excel spreadsheets."""

    def read_excel(self, path: str | Path) -> dict[str, Any]:
        """Read sheets, headers, and cell values from an Excel spreadsheet.

        Args:
            path: Path to the .xlsx file on local disk.

        Returns:
            Structured dictionary with 'sheet_names', 'sheets' (headers & rows), and 'row_counts'.
        """
        file_path = Path(path)
        if not file_path.exists():
            raise FileNotFoundError(f"Spreadsheet file not found: {file_path}")

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        sheets_data: dict[str, Any] = {}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[Any]] = []
            for row in ws.iter_rows(values_only=True):
                # Filter out completely empty rows
                if any(cell is not None for cell in row):
                    rows.append(list(row))

            headers: list[str] = []
            data_rows: list[list[Any]] = []
            if rows:
                headers = [str(c) if c is not None else "" for c in rows[0]]
                data_rows = rows[1:]

            sheets_data[sheet_name] = {
                "headers": headers,
                "rows": data_rows,
                "total_rows": len(data_rows),
            }

        return {
            "path": str(file_path),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_data,
        }

    def write_excel(
        self,
        path: str | Path,
        data: Any,
        sheet_name: str = "Sheet1",
    ) -> str:
        """Write structured tabular data to an Excel file.

        Args:
            path: Destination path for the .xlsx file.
            data: Either a dict mapping column names to lists of values, or a list of row dicts.
            sheet_name: Title of the worksheet.

        Returns:
            The resolved output path of the generated Excel file.
        """
        out_path = Path(path)
        out_path.parent.mkdir(parents=True, exist_ok=True)

        if isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame(data)

        df.to_excel(str(out_path), sheet_name=sheet_name, index=False)
        return str(out_path)

    def summary_stats(
        self, path: str | Path, sheet_name: str | None = None
    ) -> dict[str, Any]:
        """Compute summary statistics (min, max, mean, count) for numeric columns."""
        file_path = Path(path)
        if not file_path.exists():
            raise FileNotFoundError(f"Spreadsheet file not found: {file_path}")

        df = pd.read_excel(str(file_path), sheet_name=sheet_name or 0)
        numeric_df = df.select_dtypes(include=["number"])

        if numeric_df.empty:
            return {"columns": list(df.columns), "numeric_columns": [], "stats": {}}

        desc = numeric_df.describe().to_dict()
        return {
            "columns": list(df.columns),
            "numeric_columns": list(numeric_df.columns),
            "stats": desc,
        }
